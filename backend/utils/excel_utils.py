from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, date, timedelta

# ============================================================
# 상품(북두칠성1/2/3) 정의
#   A형(bdc1, bdc3): 일타수 × 구동일수 = 총타수
#   B형(bdc2)      : 일자별 작업량 D-1~D-7 합계 = 총작업량
# ============================================================
PRODUCTS = {
    'bdc1': {'title': '북 두 칠 성 1', 'label': '북두칠성1', 'run_time': '익일 구동', 'format': 'A'},
    'bdc3': {'title': '북 두 칠 성 3', 'label': '북두칠성3', 'run_time': '익일 구동', 'format': 'A'},
    'bdc2': {'title': '북두칠성 2', 'label': '북두칠성2', 'run_time': '익일 구동', 'format': 'B'},
    'bdcnav': {'title': '북두칠성 길찾기', 'label': '북두칠성 길찾기', 'run_time': '익일 구동', 'format': 'B'},
}

MAX_DAYS = 7  # D-1 ~ D-7

_thin = Side(style='thin', color='D0D0D0')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
CREAM = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
YELLOW = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
BOLD = Font(bold=True)
RED_BOLD = Font(color='FF0000', bold=True)

A_HEADERS = ['구동 시작일', '플레이스 업체명', '메인키워드', 'url (모바일)', '일타수', '구동일수', '총타수 (수정x)']
A_WIDTHS = [14, 22, 18, 40, 10, 10, 14]
B_HEADERS = ['접수일', '메인키워드', '업체명', '플레이스 링크', '시작일', '만료일']
B_WIDTHS = [13, 16, 18, 42, 13, 13]


# ---------------------------------------------------------------- helpers
def _to_date(v):
    if v is None or v == '':
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().replace(' ', '')
    for fmt in ('%Y.%m.%d', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_int(v):
    if v is None or v == '':
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


# ---------------------------------------------------------------- template
def _build_A_info(ws, product):
    ws['A1'] = '최소기간'; ws['B1'] = '7일(최소구동일자)'
    ws['A2'] = '최소 일타수'; ws['B2'] = '100타'
    ws['A3'] = '환불 불가'; ws['B3'] = 'A/S 불가'
    ws['C3'] = '빨간글씨는 수정 x (자동기입 됩니다.)'
    ws['A4'] = '구동시간'; ws['B4'] = product['run_time']
    ws['C1'] = product['title']
    for coord in ('A1', 'B1', 'C1', 'A2', 'B2', 'A3', 'B3', 'C3', 'A4', 'B4'):
        ws[coord].font = BOLD
    ws['C1'].alignment = CENTER
    for m in ('C1:G2', 'C3:G4'):
        ws.merge_cells(m)
    # 헤더 5~6행 (병합)
    for col_idx, (name, width) in enumerate(zip(A_HEADERS, A_WIDTHS), 1):
        cell = ws.cell(row=5, column=col_idx, value=name)
        cell.font = RED_BOLD if col_idx == 7 else BOLD
        cell.alignment = CENTER
        cell.border = BORDER
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = width
        ws.merge_cells(f'{col_letter}5:{col_letter}6')


def _build_B_header(ws, product, max_days=MAX_DAYS):
    """B형 헤더. max_days 만큼 D-1~D-max_days 열 생성. 반환: 총작업량 열 인덱스(1-based)"""
    max_days = max(int(max_days or MAX_DAYS), 1)
    for col_idx, (name, width) in enumerate(zip(B_HEADERS, B_WIDTHS), 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = BOLD; cell.fill = CREAM; cell.alignment = CENTER; cell.border = BORDER
        col_letter = cell.column_letter
        ws.column_dimensions[col_letter].width = width
        ws.merge_cells(f'{col_letter}1:{col_letter}2')
    # 일자별 작업량[타] 병합 (G1 ~ 마지막 일자열)
    first_day_col = 7
    last_day_col = 6 + max_days
    g = ws.cell(row=1, column=first_day_col, value='일자별 작업량[타]')
    g.font = BOLD; g.fill = CREAM; g.alignment = CENTER; g.border = BORDER
    ws.merge_cells(f'{get_column_letter(first_day_col)}1:{get_column_letter(last_day_col)}1')
    for i in range(max_days):
        c = ws.cell(row=2, column=first_day_col + i, value=f'D-{i + 1}')
        c.font = BOLD; c.fill = CREAM; c.alignment = CENTER; c.border = BORDER
        ws.column_dimensions[c.column_letter].width = 8
    total_col = last_day_col + 1
    n = ws.cell(row=1, column=total_col, value='총작업량')
    n.font = BOLD; n.fill = CREAM; n.alignment = CENTER; n.border = BORDER
    ws.merge_cells(f'{get_column_letter(total_col)}1:{get_column_letter(total_col)}2')
    ws.column_dimensions[get_column_letter(total_col)].width = 12
    return total_col


def generate_template(product_type):
    product = PRODUCTS.get(product_type, PRODUCTS['bdc1'])
    wb = Workbook()
    ws = wb.active

    if product['format'] == 'A':
        ws.title = '트래픽'
        _build_A_info(ws, product)
        # 예시 행 (7행)
        ex = [datetime(2026, 5, 30), '풋사랑정형외과의원', '부산정형외과',
              'https://m.place.naver.com/hospital/1127389017/home', 200, 7]
        for col_idx, val in enumerate(ex, 1):
            ws.cell(row=7, column=col_idx, value=val).alignment = CENTER
        tot = ws.cell(row=7, column=7, value='=E7*F7')
        tot.font = RED_BOLD; tot.alignment = CENTER
        mk = ws.cell(row=7, column=8, value='예시'); mk.fill = YELLOW
    else:
        ws.title = 'Sheet1'
        _build_B_header(ws, product)
        ex = ['2026. 6.18', '예시1', '예시2', 'https://m.place.naver.com/place/1670099706',
              '2026. 6. 19', '2026. 6. 25', 100, 250, 110, 150, 180, 130, 300, 1220]
        for col_idx, val in enumerate(ex, 1):
            ws.cell(row=3, column=col_idx, value=val).alignment = CENTER

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------- parse
def parse_campaign_excel(file_bytes, product_type, user_id, created_by):
    product = PRODUCTS.get(product_type, PRODUCTS['bdc1'])
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    results, errors = [], []

    if product['format'] == 'A':
        # 헤더 5행, 데이터 7행부터
        for row_idx, row in enumerate(rows[6:], 7):
            if not row:
                continue
            place = row[1] if len(row) > 1 else None
            keyword = row[2] if len(row) > 2 else None
            marker = row[7] if len(row) > 7 else None
            if marker == '예시':
                continue
            if not place and not keyword:
                continue
            if not place or not keyword:
                errors.append({'row': row_idx, 'error': '업체명/메인키워드 누락'})
                continue
            start = _to_date(row[0] if len(row) > 0 else None)
            daily_ta = _to_int(row[4] if len(row) > 4 else None) or 0
            run_days = _to_int(row[5] if len(row) > 5 else None) or 0
            end = start + timedelta(days=run_days) if (start and run_days) else None
            results.append({
                'user_id': user_id, 'created_by': created_by, 'product_type': product_type,
                'place_name': str(place), 'keyword_main': str(keyword),
                'place_url': str(row[3]) if len(row) > 3 and row[3] else None,
                'start_date': start, 'end_date': end,
                'daily_ta': daily_ta, 'run_days': run_days,
                'total_ta': daily_ta * run_days,
                'status': 'pending',
            })
    else:
        # 일자열 개수를 동적으로 판정: 헤더 1행에서 '총작업량' 열을 찾음 → 일자열 = index 6 ~ (총작업량-1)
        header = rows[0] if rows else ()
        total_col = None
        for i, v in enumerate(header):
            if v is not None and str(v).strip() == '총작업량':
                total_col = i
                break
        if total_col is None:
            # 폴백: 2행의 'D-n' 헤더 개수로 판정
            drow = rows[1] if len(rows) > 1 else ()
            n = sum(1 for j in range(6, len(drow)) if drow[j] is not None and str(drow[j]).startswith('D-'))
            total_col = 6 + (n if n else MAX_DAYS)
        day_cols = list(range(6, max(total_col, 7)))  # 최소 D-1 한 칸은 확보

        # 헤더 1~2행, 데이터 3행부터
        for row_idx, row in enumerate(rows[2:], 3):
            if not row:
                continue
            keyword = row[1] if len(row) > 1 else None
            place = row[2] if len(row) > 2 else None
            # 양식의 기본 예시행(예시1/예시2)만 건너뜀 — 실제 데이터는 그대로 등록
            if str(keyword) == '예시1' and str(place) == '예시2':
                continue
            if not place and not keyword:
                continue
            if not place or not keyword:
                errors.append({'row': row_idx, 'error': '업체명/메인키워드 누락'})
                continue
            days = []
            for pos, col in enumerate(day_cols):
                ta = _to_int(row[col] if len(row) > col else None)
                if ta:
                    days.append({'day_no': pos + 1, 'ta': ta})
            total = sum(d['ta'] for d in days)
            start = _to_date(row[4] if len(row) > 4 else None)
            results.append({
                'user_id': user_id, 'created_by': created_by, 'product_type': product_type,
                'place_name': str(place), 'keyword_main': str(keyword),
                'place_url': str(row[3]) if len(row) > 3 and row[3] else None,
                'intake_date': _to_date(row[0] if len(row) > 0 else None),
                'start_date': start,
                'end_date': _to_date(row[5] if len(row) > 5 else None),
                'run_days': len(day_cols),
                'total_ta': total, 'days': days,
                'status': 'pending',
            })

    return results, len(results), errors


# ---------------------------------------------------------------- export list
def _write_A_sheet(ws, product, campaigns):
    _build_A_info(ws, product)
    r = 7
    for c in campaigns:
        ws.cell(row=r, column=1, value=str(c.get('start_date') or ''))
        ws.cell(row=r, column=2, value=c.get('place_name'))
        ws.cell(row=r, column=3, value=c.get('keyword_main'))
        ws.cell(row=r, column=4, value=c.get('place_url'))
        ws.cell(row=r, column=5, value=c.get('daily_ta'))
        ws.cell(row=r, column=6, value=c.get('run_days'))
        ws.cell(row=r, column=7, value=c.get('total_ta')).font = RED_BOLD
        r += 1


def _write_B_sheet(ws, product, campaigns, days_map):
    # 캠페인들 중 최대 일수로 D열 개수 결정
    max_days = MAX_DAYS
    for c in campaigns:
        dm = days_map.get(c.get('id'), [])
        cand = max([d['day_no'] for d in dm], default=0)
        max_days = max(max_days, cand, int(c.get('run_days') or 0))
    total_col = _build_B_header(ws, product, max_days)
    r = 3
    for c in campaigns:
        ws.cell(row=r, column=1, value=str(c.get('intake_date') or ''))
        ws.cell(row=r, column=2, value=c.get('keyword_main'))
        ws.cell(row=r, column=3, value=c.get('place_name'))
        ws.cell(row=r, column=4, value=c.get('place_url'))
        ws.cell(row=r, column=5, value=str(c.get('start_date') or ''))
        ws.cell(row=r, column=6, value=str(c.get('end_date') or ''))
        day_by_no = {d['day_no']: d['ta'] for d in days_map.get(c.get('id'), [])}
        for i in range(max_days):
            ws.cell(row=r, column=7 + i, value=day_by_no.get(i + 1))
        ws.cell(row=r, column=total_col, value=c.get('total_ta'))
        r += 1


def _write_campaign_sheet(ws, product_type, campaigns, days_map):
    product = PRODUCTS.get(product_type, PRODUCTS['bdc1'])
    if product['format'] == 'A':
        _write_A_sheet(ws, product, campaigns)
    else:
        _write_B_sheet(ws, product, campaigns, days_map or {})


def export_campaigns_excel(campaigns, product_type, days_map=None):
    """단일 상품 접수양식 다운로드"""
    wb = Workbook()
    _write_campaign_sheet(wb.active, product_type, campaigns, days_map or {})
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_campaigns_multi(groups):
    """상품별 다중 시트 접수양식 다운로드. groups={product_type: (campaigns, days_map)}"""
    wb = Workbook()
    wb.remove(wb.active)
    if not groups:
        wb.create_sheet('없음')
    for pt, (campaigns, days_map) in groups.items():
        label = PRODUCTS.get(pt, {}).get('label', pt)
        ws = wb.create_sheet(str(label)[:31])
        _write_campaign_sheet(ws, pt, campaigns, days_map or {})
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ---------------------------------------------------------------- export logs
def export_logs_excel(logs, changes_map=None):
    changes_map = changes_map or {}
    wb = Workbook()
    ws = wb.active
    ws.title = '로그'

    columns = [
        ('번호', 8), ('구분', 10), ('사용자ID', 18), ('수정자', 18),
        ('캠페인번호', 12), ('상품유형', 12), ('총타수', 12), ('구동일수', 10),
        ('생성일시', 22), ('변경항목', 16), ('변경전', 30), ('변경후', 30),
    ]
    header_fill = PatternFill(start_color='1a2744', end_color='1a2744', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    old_font = Font(color='DC2626')
    new_font = Font(color='16A34A')

    for col_idx, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = width

    def base_cols(r, log):
        ws.cell(row=r, column=1, value=log.get('id'))
        ws.cell(row=r, column=2, value=log.get('type'))
        ws.cell(row=r, column=3, value=log.get('username') or log.get('user_id'))
        ws.cell(row=r, column=4, value=log.get('modified_by_username') or '-')
        ws.cell(row=r, column=5, value=log.get('campaign_id') or '-')
        ws.cell(row=r, column=6, value=PRODUCTS.get(log.get('product_type'), {}).get('label', log.get('product_type') or '-'))
        ws.cell(row=r, column=7, value=log.get('total_ta') or 0)
        ws.cell(row=r, column=8, value=log.get('period_days') or 0)
        ws.cell(row=r, column=9, value=str(log.get('created_at', '') or ''))

    row_idx = 2
    for log in logs:
        details = changes_map.get(log.get('id'), [])
        if log.get('type') == '수정' and details:
            for i, d in enumerate(details):
                if i == 0:
                    base_cols(row_idx, log)
                ws.cell(row=row_idx, column=10, value=d.get('field_label', d.get('field_name', '')))
                ws.cell(row=row_idx, column=11, value=d.get('old_value') or '-').font = old_font
                ws.cell(row=row_idx, column=12, value=d.get('new_value') or '-').font = new_font
                row_idx += 1
        else:
            base_cols(row_idx, log)
            row_idx += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
